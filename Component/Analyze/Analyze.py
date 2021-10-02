import re

from openpyxl import load_workbook

from . import Student
from . import Exam
from ..Util.File import *
from . import Result


# 모의고사 채점/분석 페이지
# 모의고사의 정답 파일과 학생 제출 파일을 읽어 채점을 진행합니다. 이후, 결과 파일들을 생성합니다.
def analyze_exam():
    exam_folder = select_exam_folder()
    exam_data = get_exam_data_from_excel(exam_folder + f'/{exam_folder.lstrip("./data")} 정답 및 배점.xlsx')
    if not exam_data:
        return
    students = get_students(exam_folder + f'/{exam_folder.lstrip("./data")} 학생 제출 답.xlsx', *exam_data)
    if not students:
        return

    exam = Exam.Exam(exam_folder, students, *exam_data)

    Result.create_result_files(exam_folder, exam)
    print("채점되었습니다!")


def get_exam_data_from_excel(exam_path):
    exam_answers = exam_answers_from_excel(exam_path)
    exam_scores = exam_scores_from_excel(exam_path)

    if exam_answers and exam_scores:
        return exam_answers, exam_scores
    return False


def exam_answers_from_excel(exam_path):
    values = load_workbook(exam_path, data_only=True)['문항 정답 및 배점'].values
    answers = [value[1] for value in values if value and type(value[1]) == int]

    if len(answers) != 20:
        print("문제 정답 개수를 다시 확인해 주세요!")
        return False
    return answers


def exam_scores_from_excel(exam_path):
    values = load_workbook(exam_path, data_only=True)['문항 정답 및 배점'].values
    scores = [value[2] for value in values if value and type(value[2]) == int]

    if len(scores) != 20:
        print("문제 배점 개수를 다시 확인해 주세요!")
        return False

    return scores


# 학생 객체를 담은 학생들 배열을 반환합니다.
def get_students(students_path, exam_answers, exam_scores):
    students = {}
    records = load_workbook(students_path, data_only=True)['제출답안'].values

    next(records)
    for name, branch, submission in records:
        if is_valid_student_data(name, branch, submission):
            student = Student.Student(name, branch, submission, exam_answers, exam_scores)
            students[student.__hash__()] = student
        else:
            print("학생 제출 답안 엑셀 파일을 다시 확인해 주세요! 잘못된 값이 들어갔습니다.")
            return False

    if not students:
        print('학생이 없습니다!')
        return False
    return students


def is_valid_student_data(name, branch, submission):
    if name and branch and submission:
        if type(name) == str and type(branch) == str and re.compile('[0-9]{20}').match(submission):
            return True
    return False
