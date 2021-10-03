import re
from openpyxl import load_workbook

from . import Student
from . import Exam
from . import Result

# 모의고사 채점/분석 페이지


# 모의고사의 정답 파일과 학생 제출 파일을 읽어 채점을 진행합니다. 이후, 결과 파일들을 생성합니다.
def analyze_exam(exam):
    exam_folder = "./data/" + exam
    exam_data = get_exam_data_from_excel(exam_folder + f'/{exam} 정답 및 배점.xlsx')
    if not exam_data:
        return 3
    students = get_students(exam_folder + f'/{exam} 학생 제출 답.xlsx', *exam_data)
    if not students:
        return 4

    exam = Exam.Exam(exam_folder, students, *exam_data)

    result = Result.create_result_files(exam_folder, exam)
    if result == 6:
        return 6
    return 100


def get_exam_data_from_excel(exam_path):
    exam_answers = exam_answers_from_excel(exam_path)
    exam_scores = exam_scores_from_excel(exam_path)

    if exam_answers and exam_scores:
        return exam_answers, exam_scores
    return False


def exam_answers_from_excel(exam_path):
    values = load_workbook(exam_path, data_only=True)['문항 정답 및 배점'].values
    next(values)
    answers = []

    for value in values:
        if not value or not (type(value[1]) == int) or not (0 < value[1] <= 5):
            return False
        answers.append(value[1])

    if len(answers) != 20:
        return False
    return answers


def exam_scores_from_excel(exam_path):
    values = load_workbook(exam_path, data_only=True)['문항 정답 및 배점'].values
    next(values)
    scores = []

    for value in values:
        if not value or not (type(value[2]) == int):
            return False
        scores.append(value[2])

    if len(scores) != 20:
        return False

    return scores


# 학생 객체를 담은 학생들 배열을 반환합니다.
def get_students(students_path, exam_answers, exam_scores):
    students = {}
    records = load_workbook(students_path, data_only=True)['제출답안'].values
    if not records:
        return False

    next(records)
    for name, branch, submission in records:
        if is_valid_student_data(name, branch, submission):
            student = Student.Student(name, branch, submission, exam_answers, exam_scores)
            students[student.__hash__()] = student
        else:
            return False

    return students


def is_valid_student_data(name, branch, submission):
    if name and branch and submission:
        if type(name) == str and type(branch) == str and type(submission) == str and re.compile('[0-9]{20}').match(submission):
            return True
    return False
