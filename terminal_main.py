import os
from openpyxl import Workbook
from openpyxl import load_workbook

# 초기 화면
def show_menu():
    print('1. 모의고사 등록 2. 모의고사 채점 3. 모의고사 열람')
    cmd = input('원하는 기능: ')
    return cmd

# ------------------------------------------------
# ------------------------------------------------
# ------------------------------------------------

# 모의고사 등록 페이지
def register_exam():
    exam_folder_path = create_exam_folder()
    exam_excel_path = create_exam_excel(exam_folder_path)
    open_exam_excel(exam_excel_path)

# 모의고사 이름을 받고, 해당 모의고사 이름으로 된 폴더 생성
def create_exam_folder():
    exam_name = input('모의고사 이름을 입력: ')
    exam_folder_path = './data/' + exam_name
    # 이미 같은 이름의 모의고사가 존재하지 않는지 확인
    if not os.path.exists(exam_folder_path):
        # 폴더명으로 사용할 수 있는지 확인
        try:
            # 폴더 만들고, 폴더의 경로를 return
            os.mkdir(exam_folder_path)
        # 오류 나오면 메세지 띄우고 이름 입력부터 다시 시작
        except:
            print('폴더명으로 쓸 수 있는 형식으로 모의고사 이름을 입력해주세요.')
            exam_folder_path = create_exam_folder()
        finally:
            return exam_folder_path
    # 오류 나오면 메세지 띄우고 이름 입력부터 다시 시작
    else:
        print('이미 존재하는 모의고사입니다.')
        print('이미 등록하신 모의고사의 정보를 수정하시려면, data 폴더 내에 있는 해당 시험의 엑셀파일에서 직접 수정하고 저장하시고 창을 닫으십시오.')
        exam_folder_path = create_exam_folder()
    return exam_folder_path

# 해당 폴더 안에 문제 정보 엑셀 파일 생성
def create_exam_excel(path):
    # 엑셀 시트 활성화
    write_wb = Workbook()
    write_ws = write_wb.create_sheet('문항 정답 및 배점')
    # 엑셀에 기본 값 넣어두기
    write_ws['A1'] = '문항'
    write_ws['B1'] = '정답'
    write_ws['C1'] = '배점'
    for question_number in range(1,21):
        cell_path = 'A' + str(question_number + 1)
        write_ws[cell_path] = question_number
    # 엑셀 파일이 저장될 경로
    exam_excel_path = str(path) + '/문항및배점.xlsx'
    # 엑셀 파일 저장 및 경로 return
    write_wb.save(exam_excel_path)
    return exam_excel_path

# 만들어진 문제 정보 엑셀파일을 띄워주는 함수
def open_exam_excel(exam_excel_path):
    if os.name == 'nt':
        os.system("start" + exam_excel_path)
    elif os.name == 'posix':
        os.system("open" + exam_excel_path)

# ------------------------------------------------
# ------------------------------------------------
# ------------------------------------------------

# 모의고사 채점 페이지
def grade_exam():
    grade_exam_path = get_grade_exam_path()
    grade_excel_path = create_grade_excel(grade_exam_path)
    open_grade_excel(grade_excel_path)

# 채점할 모의고사 이름 및 경로 받기
def get_grade_exam_path():
    # 모의고사 목록 출력
    exam_list = os.listdir('./data')
    for idx in range(len(exam_list)):
        print(str(idx+1)+ '.', exam_list[idx])
    # 채점할 모의고사 이름 받기
    grade_exam_idx = input('몇번째 모의고사를 채점하시겠습니까? ')
    grade_exam_name = exam_list[int(grade_exam_idx)-1]
    # 해당 모의고사 경로 return
    grade_exam_path = './data/' + grade_exam_name
    return grade_exam_path

# 해당 모의고사 폴더 안에 제출답안 엑셀 파일 생성
def create_grade_excel(grade_exam_path):
    # 엑셀 시트 활성화
    write_wb = Workbook()
    write_ws = write_wb.create_sheet('제출답안')
    # 엑셀에 기본 값 넣어두기
    write_ws['A1'] = '학생 이름'
    write_ws['B1'] = '지점명'
    write_ws['C1'] = '제출답안'
    # 엑셀 파일이 저장될 경로
    grade_excel_path = str(grade_exam_path) + '/제출답안.xlsx'
    # 엑셀 파일 저장 및 경로 return
    write_wb.save(grade_excel_path)
    return grade_excel_path

# 만들어진 제출답안 엑셀파일을 띄워주는 함수
def open_grade_excel(grade_excel_path):
    if os.name == 'nt':
        os.system("start" + grade_excel_path)
    elif os.name == 'posix':
        os.system("open" + grade_excel_path)

# ------------------------------------------------
# ------------------------------------------------
# ------------------------------------------------

# 모의고사 분석 페이지
def analyze_exam():
    analysis_exam_path = get_analysis_exam_path()
    data_box = get_information(analysis_exam_path)
    calculated_data_box = calculate(data_box)
    cmd = showInfo(analysis_exam_path.strip('./data/'), calculated_data_box)
    if cmd == '1':
        students_info_excel_path = writeStudentsInfo(analysis_exam_path, calculated_data_box)
        open_students_info_excel(students_info_excel_path)
    elif cmd == '2':
        exam_info_excel_path = writeExamInfo(analysis_exam_path, calculated_data_box)
        open_exam_info_excel(exam_info_excel_path)
    elif cmd == '3':
        branch_folder_path = create_branch_folder(analysis_exam_path)
        writeBranchInfo(analysis_exam_path, branch_folder_path, calculated_data_box)
        open_branch_info_excel(branch_folder_path)

# 분석할 모의고사 이름 및 경로 받기
def get_analysis_exam_path():
    # 모의고사 목록 출력
    exam_list = os.listdir('./data')
    for idx in range(len(exam_list)):
        print(str(idx+1)+ '.', exam_list[idx])
    # 채점할 모의고사 이름 받기
    analysis_exam_idx = input('몇번째 모의고사를 분석/열람하시겠습니까? ')
    analysis_exam_name = exam_list[int(analysis_exam_idx)-1]
    # 해당 모의고사 경로 return
    analysis_exam_path = './data/' + analysis_exam_name
    return analysis_exam_path

# 모의고사 관련 정보 받기
def get_information(analysis_exam_path):
    # 가져온 정보를 저장할 리스트
    data_box = []

    # 문항 및 배점 엑셀 파일 불러오기
    exam_path = analysis_exam_path + str('/문항및배점.xlsx')
    load_wb = load_workbook(exam_path, data_only=True)
    load_ws = load_wb['문항 정답 및 배점']
    # 정답 및 배점 불러오기
    answers = getAnswers(load_ws)
    scores = getScores(load_ws)

    # 제출답안 엑셀 파일 불러오기
    grade_path = analysis_exam_path + str('/제출답안.xlsx')
    load_wb = load_workbook(grade_path, data_only=True)
    load_ws = load_wb['제출답안']
    names = getNames(load_ws)
    branches = getBranches(load_ws)
    submission = getSubmission(load_ws)

    # 모은 데이터 저장하기
    data_box.append(answers)
    data_box.append(scores)
    data_box.append(names)
    data_box.append(branches)
    data_box.append(submission)

    return data_box

# 각 문항 별 정답 리스트로 저장
def getAnswers(load_ws):
    answers = []
    for answer_idx in range(2, 22):
        cell_path = 'B' + str(answer_idx)
        answer = load_ws[cell_path].value
        answers.append(answer)
    return answers

# 각 문항 별 배점 리스트로 저장
def getScores(load_ws):
    scores = []
    for score_idx in range(2, 22):
        cell_path = 'C' + str(score_idx)
        score = load_ws[cell_path].value
        scores.append(score)
    return scores


# 학생 목록 리스트로 저장
def getNames(load_ws):
    names = []
    col = load_ws['A']
    for cell in col:
        name = cell.value
        # 처음 제목과 공란 제외
        if name == '학생 이름':
            continue
        elif name == None:
            continue
        else:
            names.append(name)
    return names

# 지점 목록 리스트로 저장
def getBranches(load_ws):
    branches = []
    col = load_ws['B']
    for cell in col:
        branch = cell.value
        # 처음 제목과 공란 제외
        if branch == '지점명':
            continue
        elif branch == None:
            continue
        else:
            branches.append(branch)
    return branches

# 학생 제출답안 리스트로 저장
def getSubmission(load_ws):
    submission = []
    col = load_ws['C']
    for cell in col:
        submitted = cell.value
        submitted_list = []
        # 처음 제목과 공란 제외
        if submitted == '제출답안':
            continue
        elif submitted == None:
            continue
        else:
            # 엑셀에 연속적인 str으로 작성된 제출답안을 리스트 형태로 변환
            for idx in range(0,20):
                submitted_list.append(int(submitted[idx]))
            # 해당 제출답안들을 리스트로 저장
            submission.append(submitted_list)
    return submission

# 예시 입력값
# answers = [1, 2, 3, 4, 5, 1, 2, 3, 4, 5, 1, 2, 3, 4, 5, 1, 2, 3, 4, 5]
# names = ['이상훈', '송종빈']
# submission = [[1, 2, 3, 4, 5, 1, 2, 3, 4, 5, 1, 2, 3, 4, 5, 1, 2, 3, 4, 4], [1, 2, 3, 4, 5, 1, 2, 3, 4, 5, 1, 2, 3, 4, 5, 1, 2, 4, 4, 4]]
# scores = [5, 5, 5, 5, 5, 5, 5, 5, 5, 5, 5, 5, 5, 5, 5, 5, 5, 5, 5, 5]
# branches = ['대치', '중계']

# 학생 정보 처리하는 클래스
class StudentInfo:
    # 초기값
    def __init__(self):
        self.answers = None  # [1, 2, 3, 4, 5, 1, 2, 3, 4, 5, 1, 2, 3, 4, 5, 1, 2, 3, 4, 5]
        self.scores = None  # [5, 5, 5, 5, 5, 5, 5, 5, 5, 5, 5, 5, 5, 5, 5, 5, 5, 5, 5, 5]
        self.name = None  # '이상훈'
        self.branch = None  # '대치'
        self.submission = None  # [1, 2, 3, 4, 5, 1, 2, 3, 4, 5, 1, 2, 3, 4, 5, 1, 2, 3, 4, 4]
        self.correct_answers = None  # [1, 2, 3, 4, 5, 6, 7, 8, 9, 10, 11, 12, 13, 14, 15, 16, 17, 18, 19]
        self.wrong_answers = None  # [20]
        self.myScore = 0  # 45

    # 이름 받아오기
    def getName(self, student_name):
        self.name = student_name

    # 지점 받아오기
    def getBranch(self, branch):
        self.branch = branch

    # 맞춘 답안 리스트 구하기
    def getCorrectAnswers(self, answers, submission):
        self.answers = answers
        self.submission = submission
        self.correct_answers = []
        # 1번부터 20번까지 맞은 문제의 문제 번호를 리스트에 저장
        for idx in range(0,20):
            if self.answers[idx] == self.submission[idx]:
                self.correct_answers.append(idx+1)
            else:
                continue

    # 틀린 답안 리스트 구하기
    def getWrongAnswers(self, answers, submission):
        self.answers = answers
        self.submission = submission
        self.wrong_answers = []
        # 1번부터 20번까지 틀린 문제의 문제 번호를 리스트에 저장
        for idx in range(0, 20):
            if self.answers[idx] != self.submission[idx]:
                self.wrong_answers.append(idx+1)
            else:
                continue

    # 점수 구하기
    def getScore(self, correct_answers, scores):
        self.scores = scores
        self.correct_answers = correct_answers
        # 맞은 문항의 배점을 더해가는 방식
        for idx in correct_answers:
            self.myScore += self.scores[idx - 1]

# 시험 정보 처리하는 클래스
class ExamInfo:
    # 초기값
    def __init__(self):
        self.score_list = {}  # [('이상훈', 95), ('송종빈', 90)]
        self.wrong_answers_list = []  # [20, 18, 20]
        self.score_sum = None  # 185
        self.studnet_number = 0  # 2
        self.score_sum = 0  # 185
        self.score_average = 0  # 92.5
        self.wrong_answers_numbers = []  # [0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 1, 0, 2]
        self.wrong_answers_dictionary = {}  # [(20, 2), (18, 1), (1, 0), (2, 0), (3, 0), (4, 0), (5, 0), (6, 0), (7, 0), (8, 0), (9, 0), (10, 0), (11, 0), (12, 0), (13, 0), (14, 0), (15, 0), (16, 0), (17, 0), (19, 0)]
        self.wrong_answers_dictionary_sort = {}  # {20: 2, 18: 1, 1: 0, 2: 0, 3: 0, 4: 0, 5: 0, 6: 0, 7: 0, 8: 0, 9: 0, 10: 0, 11: 0, 12: 0, 13: 0, 14: 0, 15: 0, 16: 0, 17: 0, 19: 0}
        self.top5_wrong_answers = []  # [20, 18, 1, 2, 3]
        self.score_list_name_sort = []  # ['이상훈', '송종빈']
        self.score_list_name_dictionary = {}  # {'이상훈': 1, '송종빈': 2}
        self.score_list_scores_sort = []  # [95, 90]
        self.cut_lines = []  # [95, 95, 95, 95, 90, 90, 90, 90]
        self.cut_lines_proportion = [0.04, 0.11, 0.23, 0.40, 0.60, 0.77, 0.89, 0.96]
        self.correct_answers_numbers = []  # [2, 2, 2, 2, 2, 2, 2, 2, 2, 2, 2, 2, 2, 2, 2, 2, 2, 1, 2, 0]
        self.correct_answers_percentage = []  # ['100.0%', '100.0%', '100.0%', '100.0%', '100.0%', '100.0%', '100.0%', '100.0%', '100.0%', '100.0%', '100.0%', '100.0%', '100.0%', '100.0%', '100.0%', '100.0%', '100.0%', '50.0%', '100.0%', '0.0%']
        self.chosen_answers = {}  # {1: [2, 0, 0, 0, 0], 2: [0, 2, 0, 0, 0], 3: [0, 0, 2, 0, 0], 4: [0, 0, 0, 2, 0], 5: [0, 0, 0, 0, 2], 6: [2, 0, 0, 0, 0], 7: [0, 2, 0, 0, 0], 8: [0, 0, 2, 0, 0], 9: [0, 0, 0, 2, 0], 10: [0, 0, 0, 0, 2], 11: [2, 0, 0, 0, 0], 12: [0, 2, 0, 0, 0], 13: [0, 0, 2, 0, 0], 14: [0, 0, 0, 2, 0], 15: [0, 0, 0, 0, 2], 16: [2, 0, 0, 0, 0], 17: [0, 2, 0, 0, 0], 18: [0, 0, 1, 1, 0], 19: [0, 0, 0, 2, 0], 20: [0, 0, 0, 2, 0]}

    # 학생 수 받아오기
    def getStudentNumber(self, student_number):
        self.studnet_number = student_number

    # 총점 구하기
    def getScoreSum(self, score):
        self.score_sum += score

    # 학생 별 점수 구하기
    def getScoreList(self, name, score):
        self.score_list[name] = score

    # 틀린 답안 리스트 만들기 (전체 학생)
    def getWrongAnswersList(self, wrong_answers):
        for wrong_answer in wrong_answers:
            self.wrong_answers_list.append(wrong_answer)

    # 가장 많이 틀린 답안 찾기
    def getFrequentWrongAnswers(self):
        # 1번부터 20번까지 각 번호별로 몇번 틀렸는지 개수 세기
        for idx in range(1, 21):
            self.wrong_answers_numbers.append(self.wrong_answers_list.count(idx))

        # 각 번호별 틀린 개수를 딕셔너리로 저장
        for idx in range(0,20):
            self.wrong_answers_dictionary[idx+1] = self.wrong_answers_numbers[idx]

        # value 값을 바탕으로 내림차순 정렬
        self.wrong_answers_dictionary = sorted(self.wrong_answers_dictionary.items(), reverse=True, key=lambda item: item[1])
        for items in self.wrong_answers_dictionary:
            self.wrong_answers_dictionary_sort[items[0]] = items[1]

        # 가장 많이 틀린 순서대로 문항 번호 담기는 리스트 생성
        for items in self.wrong_answers_dictionary:
            self.top5_wrong_answers.append(items[0])

        # 가장 많이 틀린 5개까지 자름
        self.top5_wrong_answers = self.top5_wrong_answers[:5]

    # 평균 값 계산
    def getAverage(self):
        self.score_average = self.score_sum / self.studnet_number

    # 학생별 등수 정리
    def getNamesRanks(self):
        # 점수 리스트를 value 값을 기준으로 내림차순으로 정렬
        self.score_list = sorted(self.score_list.items(), reverse=True, key=lambda item: item[1])

        # 이를 이용해 가장 높은 등수부터 학생 나열
        for items in self.score_list:
            self.score_list_name_sort.append(items[0])

        # 각 학생(key)별 등수(value)를 담고 있는 딕셔너리 생성
        for idx in range(len(self.score_list_name_sort)):
            self.score_list_name_dictionary[self.score_list_name_sort[idx]] = idx + 1

    # 등수 별 점수 정리
    def getScoresRanks(self):
        # 내림차순으로 정렬되어있는 등수 리스트
        for items in self.score_list:
            self.score_list_scores_sort.append(items[1])

    # 커트라인 정리
    def getCutLines(self):
        # 각 등급컷에 해당하는 등수 구하기
        for proportion in self.cut_lines_proportion:
            cut_person = int(self.studnet_number * proportion)
            # 해당 등수의 점수 리스트에 넣기
            self.cut_lines.append(self.score_list_scores_sort[cut_person])

    # 각 문항별로 정답률 구하기
    def getCorrectRatio(self):
        # 각 문항별 맞춘 학생 수 구하기
        for idx in range(0, 20):
            self.correct_answers_numbers.append(self.studnet_number - self.wrong_answers_numbers[idx])
        # 각 문항별 정답률 구하기
        for idx in range(0, 20):
            self.correct_answers_percentage.append(str((self.correct_answers_numbers[idx] / self.studnet_number) * 100) + '%')

    # 각 문항별로 선택된 선지 순위 비어있는 틀 만들기
    def createChosenAnswers(self):
        for idx in range(1, 21):
            self.chosen_answers[idx] = [0, 0, 0, 0, 0]

    # 각 문항별로 선택된 선지 순위
    def getChosenAnswers(self, submitted_answers):
        for idx in range(len(submitted_answers)):
            chosen_answer = submitted_answers[idx]
            self.chosen_answers[idx+1][chosen_answer-1] += 1

# 필요한 값들 계산
def calculate(data_box):

    # 데이터 박스 풀기
    answers = data_box[0]
    scores = data_box[1]
    names = data_box[2]
    branches = data_box[3]
    submission = data_box[4]

    # 계산한 데이터 담을 데이터 박스 준비
    students_data_box = []
    exam_data_box = []
    calculated_data_box = []

    # 시험 정보 처리 클래스 생성
    exam = ExamInfo()
    exam.createChosenAnswers()

    # 각 학생별로 계산
    idx = 0
    for name in names:
        # st라는 이름으로 학생 정보 처리 클래스 생성
        st = StudentInfo()
        # 이름, 지점 입력
        st.getName(name)
        st.getBranch(branches[idx])
        # 정답, 오답, 점수 계산
        st.getCorrectAnswers(answers, submission[idx])
        st.getWrongAnswers(answers, submission[idx])
        st.getScore(st.correct_answers, scores)

        # 해당 학생의 데이터 박스
        student_data_box = []

        # 각 정보를 모두 학생 정보 데이터 박스에 담음
        student_data_box.append(st.name)  # student[0] = 학생 이름
        student_data_box.append(st.branch)  # student[1] = 지점명
        student_data_box.append(st.myScore)  # student[2] = 학생 점수
        student_data_box.append(st.answers)  # student[3] = 시험 정답
        student_data_box.append(st.submission)  # student[4] = 제출 답안

        # 전체 학생의 데이터 박스에 리스트 형태로 넣음
        students_data_box.append(student_data_box)

        # 학생들의 총점 계산, 학생별 점수 리스트 정리, 틀린 문제 리스트 정리
        exam.getScoreSum(st.myScore)
        exam.getScoreList(st.name, st.myScore)
        exam.getWrongAnswersList(st.wrong_answers)
        exam.getChosenAnswers(submission[idx])

        idx += 1

    # 여러 값들 입력, 정리, 계산
    exam.getStudentNumber(len(names))  # 응시생 수 입력
    exam.getFrequentWrongAnswers()  # 가장 많이 틀린 5개 문항 정리
    exam.getAverage()  # 평균값 계산
    exam.getNamesRanks()  # 등수 계산 및 이름:등수 형태의 딕셔너리 생성
    exam.getScoresRanks()  # 내림차순으로 나열된 학생들의 점수
    exam.getCutLines()  # 각 등급별 커트라인 계산
    exam.getCorrectRatio()  # 정답률 계산

    # 각 정보를 모두 시험 정보 데이터 박스에 담음
    exam_data_box.append(exam.studnet_number)  # exam[0] = 학생 수
    exam_data_box.append(exam.top5_wrong_answers)  # exam[1] = 오답 top 5 리스트
    exam_data_box.append(exam.score_average)  # exam[2] = 평균 점수
    exam_data_box.append(exam.score_list_name_dictionary)  # exam[3] = {학생이름:점수} 딕셔너리
    exam_data_box.append(exam.cut_lines)  # exam[4] = 커트라인 점수 리스트
    exam_data_box.append(exam.correct_answers_percentage)  # exam[5] = 각 문항 정답률 리스트
    exam_data_box.append(exam.chosen_answers)  # exam[6] = 고른 선지 리스트

    # 학생 정보 데이터 박스와 시험 정보 데이터 박스를 전체 계산된 데이터 박스에 담음
    calculated_data_box.append(students_data_box)
    calculated_data_box.append(exam_data_box)

    return calculated_data_box

# 모의고사 정보 보여줌
def showInfo(exam_name, calculated_data_box):
    print(exam_name, '(응시생: %d명)'%(calculated_data_box[1][0]))  # 시험 이름, 응시생 수
    print('평균 점수:', calculated_data_box[1][2])  # 평균 점수
    print('오답 1~5순위:', calculated_data_box[1][1])  # 오답 1~5 순위
    print('1~9등급 커트라인')  # 각 등급 커트라인
    for idx in range(1,9):
        print(str(idx) + '등급: ' + str(calculated_data_box[1][4][idx-1]))
    # 다음 명령창 띄우기
    print('1. 학생 전체 점수 열람 2. 각 문항별 분석 3. 학생 성적표 열람')
    cmd = input('열람할 자료를 고르세요: ')
    return cmd

# 학생 전체 점수 열람 엑셀 생성
def writeStudentsInfo(analysis_exam_path, calculated_data_box):
    # 엑셀 시트 활성화
    write_wb = Workbook()
    write_ws = write_wb.create_sheet('학생 전체 점수표')
    # 엑셀에 기본 값 넣어두기
    write_ws['A1'] = '이름'
    write_ws['B1'] = '점수'
    write_ws['C1'] = '등급'
    write_ws['D1'] = '등수'

    # 등급컷 비율
    cut_lines_proportion = [0.04, 0.11, 0.23, 0.40, 0.60, 0.77, 0.89, 0.96, 1.00]

    # 각 학생별로 필요 정보 구하고 엑셀에 넣기
    for idx in range(len(calculated_data_box[0])):
        name = calculated_data_box[0][idx][0]  # 학생 이름
        score = calculated_data_box[0][idx][2]  # 학생 점수
        rank = calculated_data_box[1][3][name]  # 학생 등수
        students_number = len(calculated_data_box[0])  # 학생수
        # 학생 등급 구하기 (등수/학생수과 등급컷 비율 대소비교를 통해)
        score_class = 0
        for proportion in range(len(cut_lines_proportion)):
            if rank / students_number <= cut_lines_proportion[proportion]:
                score_class = proportion + 1
                break
        # 셀 값의 위치에 맞춰서 해당 값들을 넣음
        cell_path = str(idx + 2)
        write_ws['A'+cell_path] = name
        write_ws['B'+cell_path] = score
        write_ws['C'+cell_path] = score_class
        write_ws['D'+cell_path] = rank

    # 엑셀 파일이 저장될 경로
    students_info_excel_path = str(analysis_exam_path) + '/학생 전체 점수 열람.xlsx'
    # 엑셀 파일 저장 및 경로 return
    write_wb.save(students_info_excel_path)
    return students_info_excel_path


# 시험 정보 열람 엑셀 생성
def writeExamInfo(analysis_exam_path, calculated_data_box):
    # 엑셀 시트 활성화
    write_wb = Workbook()
    write_ws = write_wb.create_sheet('학생 전체 점수표')
    # 엑셀에 기본 값 넣어두기
    write_ws['A1'] = '문항'
    write_ws['B1'] = '정답'
    write_ws['C1'] = '정답률'
    write_ws['D1'] = '선지선택비율'

    # 문항 번호 적기
    for idx in range(1, 21):
        cell_path = 'A' + str(idx + 1)
        write_ws[cell_path] = idx

    # 정답 적기
    answers = calculated_data_box[0][0][4]
    for idx in range(1, 21):
        cell_path = 'B' + str(idx + 1)
        write_ws[cell_path] = answers[idx - 1]

    # 정답률 적기
    correct_answers_percentage = calculated_data_box[1][5]
    for idx in range(1, 21):
        cell_path = 'C' + str(idx + 1)
        write_ws[cell_path] = correct_answers_percentage[idx - 1]

    # 선지선택순위 적기
    chosen_answers = calculated_data_box[1][6]
    student_number = len(calculated_data_box[0])
    for idx in range(1, 21):
        cell_path = 'D' + str(idx + 1)
        chosen_answers_ratio = ''
        for idx2 in range(1, 6):
            chosen_answers_ratio += str(idx2)
            chosen_answers_ratio += ': '
            chosen_answers_ratio += str((chosen_answers[idx][idx2 - 1] / student_number) * 100)
            chosen_answers_ratio += '%'
            if idx2 != 5:
                chosen_answers_ratio += '   /   '
        write_ws[cell_path] = chosen_answers_ratio

    # 엑셀 파일이 저장될 경로
    exam_info_excel_path = str(analysis_exam_path) + '/시험 정보 열람.xlsx'
    # 엑셀 파일 저장 및 경로 return
    write_wb.save(exam_info_excel_path)
    return exam_info_excel_path

# 지점의 폴더를 만드는 기능
def create_branch_folder(analysis_exam_path):
    branch_name = input('지점명을 입력: ')
    branch_folder_path = analysis_exam_path + '/' + branch_name
    # 이미 같은 이름의 지점이 존재하지 않는지 확인
    if not os.path.exists(branch_folder_path):
        # 폴더명으로 사용할 수 있는지 확인
        try:
            # 폴더 만들고, 폴더의 경로를 return
            os.mkdir(branch_folder_path)
            return branch_folder_path
        # 오류 나오면 메세지 띄우고 이름 입력부터 다시 시작
        except:
            print('폴더명으로 쓸 수 있는 형식으로 지점 이름을 입력해주세요.')
            create_branch_folder(analysis_exam_path)
    # 오류 나오면 메세지 띄우고 이름 입력부터 다시 시작
    else:
        print('이미 존재하는 지점입니다.')
        print('이미 등록하신 지점의 정보를 수정하시려면, data 폴더 내에 있는 해당 모의고사 폴더 내에 있는 지점 폴더를 삭제하십시오.')
        create_branch_folder(analysis_exam_path)

    return branch_folder_path

# 지점별 점수 열람 엑셀 생성
def writeBranchInfo(analysis_exam_path, branch_folder_path, calculated_data_box):
    student_number = len(calculated_data_box[0])
    for idx in range(student_number):
        # 엑셀 시트 활성화
        write_wb = Workbook()
        write_ws = write_wb.create_sheet('해당 학생 점수표')

        # calculated_data_box에서 정보 정리
        cut_lines_proportion = [0.04, 0.11, 0.23, 0.40, 0.60, 0.77, 0.89, 0.96, 1.00]
        exam_name = analysis_exam_path.strip('./data/')
        name = calculated_data_box[0][idx][0]
        score = calculated_data_box[0][idx][2]
        rank = calculated_data_box[1][3][name]

        # 등급 계산
        score_class = 0
        for proportion in range(len(cut_lines_proportion)):
            if rank / student_number <= cut_lines_proportion[proportion]:
                score_class = proportion + 1
                break

        # 엑셀에 기본 정보 넣기
        write_ws['A1'] = '모의고사 이름'
        write_ws['B1'] = exam_name
        write_ws['D1'] = '학생 이름'
        write_ws['E1'] = name
        write_ws['G1'] = '점수'
        write_ws['H1'] = score
        write_ws['J1'] = '등수'
        write_ws['K1'] = rank
        write_ws['M1'] = '등급'
        write_ws['N1'] = score_class

        alphabet = 'ABCDEFGHIJKLMNOPQRSTUVWXYZ'

        # 기본 정보 넣어두기 - 문항 번호, 정답, 제출답, 정답률
        for idx2 in range(1,21):
            cell_path = alphabet[idx2+1] + str(3)
            write_ws[cell_path] = idx2
        write_ws['A3'] = '문항번호'
        write_ws['A4'] = '정답'
        write_ws['A5'] = '제출답'
        write_ws['A6'] = '정답률'

        # 정답 넣기
        answers = calculated_data_box[0][0][3]
        for idx2 in range(1,21):
            cell_path = alphabet[idx2+1] + str(4)
            write_ws[cell_path] = answers[idx2-1]

        # 제출답 넣기
        submission = calculated_data_box[0][idx][4]
        for idx2 in range(1,21):
            cell_path = alphabet[idx2+1] + str(5)
            write_ws[cell_path] = submission[idx2-1]

        # 정답률 넣기
        correct_answers_percentage = calculated_data_box[1][5]
        for idx2 in range(1,21):
            cell_path = alphabet[idx2+1] + str(6)
            write_ws[cell_path] = correct_answers_percentage[idx2-1]

        # 엑셀 파일이 저장될 경로
        branch_info_excel_path = str(branch_folder_path) + '/' + name + '.xlsx'
        # 엑셀 파일 저장 및 경로 return
        write_wb.save(branch_info_excel_path)

# 만들어진 학생정보열람 엑셀파일을 띄워주는 함수
def open_students_info_excel(students_info_excel_path):
    if os.name == 'nt':
        os.system("start" + students_info_excel_path)
    elif os.name == 'posix':
        os.system("open" + students_info_excel_path)

# 만들어진 시험정보열람 엑셀파일을 띄워주는 함수
def open_exam_info_excel(exam_info_excel_path):
    if os.name == 'nt':
        os.system("start" + exam_info_excel_path)
    elif os.name == 'posix':
        os.system("open" + exam_info_excel_path)

# 만들어진 지점 폴더를 띄워주는 함수
def open_branch_info_excel(branch_folder_path):
    if os.name == 'nt':
        os.system("start" + branch_folder_path)
    elif os.name == 'posix':
        os.system("open" + branch_folder_path)

# ------------------------------------------------
# ------------------------------------------------
# ------------------------------------------------

cmd = show_menu()
if cmd == '1':
    register_exam()
elif cmd == '2':
    grade_exam()
elif cmd == '3':
    analyze_exam()
